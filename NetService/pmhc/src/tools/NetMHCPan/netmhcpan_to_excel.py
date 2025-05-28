import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path

def save_excel(output:str, output_dir:str, output_filename:str):
    # 增强正则表达式（允许最后四列部分缺失）
    table_pattern = re.compile(r"(\d+)\s+([^\s]+)\s+([A-Z*-]+)\s+([A-Z*-]+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+([A-Z*-]+)\s+([^\s]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([<= WS B]*)")
    
    # 完整17列定义
    columns = ["Pos", "MHC", "Peptide", "Core", "Of", "Gp", "Gl", "Ip", "Il", "Icore",
              "Identity", "Score_EL", "%Rank_EL", "Score_BA", "%Rank_BA", "Aff(nM)", "BindLevel"]

    # 分割不同蛋白结果块
    blocks = re.split(r"-{100,}", output)
    
    all_data = []
    for block in blocks:
        # 处理数据行
        matches = table_pattern.findall(block)
        if matches:
            # 补齐缺失字段
            processed_rows = []
            for m in matches:
                row = list(m[:12])  # 前12个固定字段
                # 处理最后5个可选字段
                row += list(m[12:]) if len(m) > 12 else ['']*5
                # 确保总长度17
                row += ['']*(17-len(row))
                processed_rows.append(row)
            all_data.extend(processed_rows)
        
        # 提取统计信息
        summary_match = re.search(r"Protein .+?\. Allele .+?\. Number of high binders \d+\. Number of weak binders \d+\. Number of peptides \d+", block)
        if summary_match:
            # 创建统计行（17列结构）
            summary_row = [summary_match.group()] + ['']*16
            all_data.append(summary_row)
    
    # 创建DataFrame
    df = pd.DataFrame(all_data, columns=columns)

    # 写入Excel
    output_path = Path(output_dir) / output_filename
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Results", index=False)
        
        # 合并统计行单元格
        workbook = writer.book
        worksheet = writer.sheets["Results"]
        for idx in df.index[df['Pos'].str.contains('Protein', na=False)]:
            excel_row = idx + 2  # 转换为Excel行号
            worksheet.merge_cells(
                start_row=excel_row,
                start_column=1,
                end_row=excel_row,
                end_column=17  # 合并全部17列
            )
            cell = worksheet.cell(row=excel_row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    return output_path