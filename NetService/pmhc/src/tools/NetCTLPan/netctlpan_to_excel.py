import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path

def save_excel(output: str, output_dir: str, output_filename: str):
    # 定义表头
    columns = ["N", "Sequence Name", "Allele", "Peptide", "MHC", "TAP", "Cle", "Comb", "%Rank"]

    # 分块（每个等位基因一块）
    blocks = re.split(r"-{20,}", output)

    all_data = []
    for block in blocks:
        # 匹配数据行（以数字开头，字段用空格分隔，允许最后一列缺失）
        data_lines = re.findall(
            r"^\s*(\d+)\s+([^\s]+)\s+([^\s]+)\s+([^\s]+)\s+([\-\d.]+)\s+([\-\d.]+)\s+([\-\d.]+)\s+([\-\d.]+)\s+([\-\d.]+)?\s*$",
            block, re.MULTILINE
        )
        for line in data_lines:
            row = list(line)
            # %Rank 可能为空，补空
            if len(row) < 9:
                row += [''] * (9 - len(row))
            all_data.append(row)

        # 匹配统计行
        summary_match = re.search(r"Number of MHC ligands.+?protein.+", block, re.IGNORECASE)
        if summary_match:
            summary_row = [summary_match.group()] + [''] * (len(columns) - 1)
            all_data.append(summary_row)

    # 写入 DataFrame
    df = pd.DataFrame(all_data, columns=columns)

    # 写入 Excel 并合并统计行
    output_path = Path(output_dir) / output_filename
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Results", index=False)
        worksheet = writer.sheets["Results"]
        for idx in df.index[df['N'].str.contains('Number of MHC ligands', na=False)]:
            excel_row = idx + 2  # Excel行号
            worksheet.merge_cells(
                start_row=excel_row,
                start_column=1,
                end_row=excel_row,
                end_column=len(columns)
            )
            cell = worksheet.cell(row=excel_row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    return output_path
