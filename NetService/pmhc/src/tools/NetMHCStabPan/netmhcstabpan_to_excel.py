import re
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path

def save_excel(output:str,output_dir:str,output_filename:str):
    table_pattern = re.compile(
        r"^\s*(\d+)\s+"      # 第1组：pos（数字）
        r"([^\s]+)\s+"      # 第2组：HLA（非空白字符）
        r"([A-Z]+)\s+"       # 第3组：peptide（大写字母）
        r"([^\s]+)\s+"       # 第4组：Identity（非空白字符）
        r"([\d.]+)\s+"       # 第5组：Pred（数字或点）
        r"([\d.]+)\s+"       # 第6组：Thalf(h)（数字或点）
        r"([\d.]+)\s*"       # 第7组：%Rank_Stab（数字或点）
        r"([<= WS B]*)"        # 第8组：BindLevel（可选，可能包含 <= WS B 等）
    , flags=re.MULTILINE)
    matches = table_pattern.findall(output)

    # 将匹配的数据转换为 DataFrame
    columns = ["Pos", "HLA", "peptide", "Identity", "Pred", "Thalf(h)", "%Rank_Stab", "BindLevel"]
    df = pd.DataFrame(matches, columns=columns)

    # 提取包含 Allele 的整行数据
    summary_pattern = re.compile(r".*Allele\s+[^\s]+.*")
    summary_match = summary_pattern.findall(output)

    # 如果找到匹配的统计信息，将其添加到 Results 表的最后一行
    if summary_match:
        # 创建一个新行，前 8列合并为一个单元格
        summary_row = [summary_match[0]] + [""] * (len(columns) - 1)  # 第一列是 summary_match[0]，其余列为空

        # 将 summary_row 添加到 DataFrame 的最后一行
        df.loc[len(df)] = summary_row
        
    output_path= Path(output_dir) / output_filename

    # 写入 Excel 文件
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Results", index=False)

        # 获取 workbook 和 worksheet 对象
        workbook = writer.book
        worksheet = writer.sheets["Results"]

        # 如果存在 summary_match，合并最后一行的前 8列并居中
        if summary_match:
            # 合并最后一行的前 8 列
            worksheet.merge_cells(start_row=len(df)+1, start_column=1, end_row=len(df)+1, end_column=8)

            # 设置合并后的单元格内容居中
            cell = worksheet.cell(row=len(df)+1, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 保存文件
    workbook.save(output_path)


    