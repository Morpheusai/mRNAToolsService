import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path

def save_excel(output: str, output_dir: str, output_filename: str):
    # 定义表头
    columns = ["Pos", "AA", "C", "score", "Ident"]

    # 分块（每个蛋白一块）
    blocks = re.split(r"-{20,}", output)

    all_data = []
    for block in blocks:
        # 匹配数据行（以数字开头，字段用空格分隔）
        data_lines = re.findall(
            r"^\s*(\d+)\s+([A-Z])\s+([.S])\s+([\d.]+)\s+([^\s]+)\s*$",
            block, re.MULTILINE
        )
        for line in data_lines:
            row = list(line)
            all_data.append(row)

        # 匹配统计行
        summary_match = re.search(r"Number of cleavage sites.+", block, re.IGNORECASE)
        if summary_match:
            summary_row = [summary_match.group()] + [''] * (len(columns) - 1)
            all_data.append(summary_row)

    # 写入 DataFrame
    df = pd.DataFrame(all_data, columns=columns)

    # 写入 Excel 并合并统计行
    output_path = Path(output_dir) / output_filename
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Results", index=False)
        worksheet = writer.sheets["Results"]
        for idx in df.index[df['Pos'].astype(str).str.contains('Number of cleavage sites', na=False)]:
            excel_row = idx + 2  # Excel行号
            worksheet.merge_cells(
                start_row=excel_row,
                start_column=1,
                end_row=excel_row,
                end_column=len(columns)
            )
            cell = worksheet.cell(row=excel_row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    return output_path